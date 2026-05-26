from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook

from data_sources import master_path as _active_master_path

# Backwards-compat constant — prefer `_active_master_path()` inside functions
# so Streamlit source-switching works.
MASTER_PATH = _active_master_path()

GymType = Literal["Ultra", "Premium"]

COL_TAG = 2
COL_MEMBERSHIP_TYPE = 4
COL_DOP = 8
COL_CASH = 11
COL_ONLINE = 12
COL_REFUND = 13
COL_BALANCE = 14
COL_DISCOUNT_AMT = 17
COL_TOTAL = 18
COL_PT = 20
COL_PAYMENT_MODE = 22
COL_CLIENT_NAME = 23
COL_LOCATION = 24


@dataclass
class ClientAggregate:
    name: str
    gym_type: str
    transaction_count: int = 0
    total_paid: float = 0.0
    cash_paid: float = 0.0
    online_paid: float = 0.0
    discount_given: float = 0.0
    refund: float = 0.0
    outstanding: float = 0.0
    locations: set[str] = field(default_factory=set)
    membership_types: set[str] = field(default_factory=set)
    payment_modes: set[str] = field(default_factory=set)
    first_seen: datetime | None = None
    last_seen: datetime | None = None
    is_pt: bool = False


@dataclass
class SegmentBreakdown:
    by_payment_mode: dict[str, dict[str, float]] = field(default_factory=dict)
    by_membership_type: dict[str, dict[str, float]] = field(default_factory=dict)
    by_location: dict[str, dict[str, float]] = field(default_factory=dict)
    pt_clients: list[ClientAggregate] = field(default_factory=list)


@dataclass
class MemberAggregates:
    total_members: dict[str, int] = field(default_factory=dict)
    total_active: dict[str, int] = field(default_factory=dict)
    total_revenue: dict[str, float] = field(default_factory=dict)
    avg_rev_per_member: dict[str, float] = field(default_factory=dict)

    new_count_by_type: dict[str, int] = field(default_factory=dict)
    existing_count_by_type: dict[str, int] = field(default_factory=dict)
    new_rev_by_type: dict[str, float] = field(default_factory=dict)
    existing_rev_by_type: dict[str, float] = field(default_factory=dict)

    monthly_new_count: dict[str, dict[str, int]] = field(default_factory=dict)
    monthly_new_revenue: dict[str, dict[str, float]] = field(default_factory=dict)
    monthly_existing_count: dict[str, dict[str, int]] = field(default_factory=dict)
    monthly_existing_revenue: dict[str, dict[str, float]] = field(default_factory=dict)

    overall_existing_members: int = 0
    overall_new_members: int = 0
    overall_existing_revenue: float = 0.0
    overall_new_revenue: float = 0.0


def _month_label(d: datetime) -> str:
    return d.strftime("%b %Y")


def _norm(v) -> str:
    return str(v).strip() if v not in (None, "") else ""


def _num(v) -> float:
    if v in (None, ""):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def _classify_tag(s) -> str | None:
    if not s:
        return None
    s = str(s).strip().lower()
    if "new" in s:
        return "New"
    if "existing" in s:
        return "Existing"
    return None


def read_membership(path: Path | None = None) -> MemberAggregates:
    if path is None:
        path = _active_master_path()
    wb = load_workbook(path, data_only=True)
    agg = MemberAggregates()

    if "Membership Summary" in wb.sheetnames:
        ms = wb["Membership Summary"]
        for r in range(4, min(ms.max_row + 1, 13)):
            label = ms.cell(r, 1).value
            if label in ("Ultra", "Premium"):
                agg.total_members[label] = ms.cell(r, 2).value or 0
                agg.total_active[label] = ms.cell(r, 3).value or 0
                agg.total_revenue[label] = ms.cell(r, 8).value or 0
                agg.avg_rev_per_member[label] = ms.cell(r, 9).value or 0
            elif label == "Existing":
                agg.overall_existing_members = ms.cell(r, 2).value or 0
                agg.overall_existing_revenue = ms.cell(r, 6).value or 0
            elif label == "New":
                agg.overall_new_members = ms.cell(r, 2).value or 0
                agg.overall_new_revenue = ms.cell(r, 6).value or 0

    for gym_type in ("Ultra", "Premium"):
        agg.new_count_by_type[gym_type] = 0
        agg.existing_count_by_type[gym_type] = 0
        agg.new_rev_by_type[gym_type] = 0.0
        agg.existing_rev_by_type[gym_type] = 0.0
        agg.monthly_new_count[gym_type] = {}
        agg.monthly_new_revenue[gym_type] = {}
        agg.monthly_existing_count[gym_type] = {}
        agg.monthly_existing_revenue[gym_type] = {}

        if gym_type not in wb.sheetnames:
            continue
        ws = wb[gym_type]
        for r in range(2, ws.max_row + 1):
            tag = _classify_tag(ws.cell(r, COL_TAG).value)
            dop = ws.cell(r, COL_DOP).value
            amount = ws.cell(r, COL_TOTAL).value or 0
            if tag is None or not isinstance(dop, datetime):
                continue
            month = _month_label(dop)
            if tag == "New":
                agg.new_count_by_type[gym_type] += 1
                agg.new_rev_by_type[gym_type] += amount
                agg.monthly_new_count[gym_type][month] = agg.monthly_new_count[gym_type].get(month, 0) + 1
                agg.monthly_new_revenue[gym_type][month] = agg.monthly_new_revenue[gym_type].get(month, 0) + amount
            else:
                agg.existing_count_by_type[gym_type] += 1
                agg.existing_rev_by_type[gym_type] += amount
                agg.monthly_existing_count[gym_type][month] = agg.monthly_existing_count[gym_type].get(month, 0) + 1
                agg.monthly_existing_revenue[gym_type][month] = agg.monthly_existing_revenue[gym_type].get(month, 0) + amount

    return agg


def read_client_aggregates(path: Path | None = None) -> tuple[dict[str, ClientAggregate], SegmentBreakdown]:
    """Aggregate raw Ultra+Premium rows into per-client totals + segment breakdowns."""
    if path is None:
        path = _active_master_path()
    wb = load_workbook(path, data_only=True)
    clients: dict[str, ClientAggregate] = {}
    seg = SegmentBreakdown()

    for gym_type in ("Ultra", "Premium"):
        seg.by_payment_mode.setdefault(gym_type, {})
        seg.by_membership_type.setdefault(gym_type, {})
        seg.by_location.setdefault(gym_type, {})

        if gym_type not in wb.sheetnames:
            continue
        ws = wb[gym_type]

        for r in range(2, ws.max_row + 1):
            name = _norm(ws.cell(r, COL_CLIENT_NAME).value)
            if not name:
                continue
            dop = ws.cell(r, COL_DOP).value
            if not isinstance(dop, datetime):
                continue

            amount = _num(ws.cell(r, COL_TOTAL).value)
            cash = _num(ws.cell(r, COL_CASH).value)
            online = _num(ws.cell(r, COL_ONLINE).value)
            refund = _num(ws.cell(r, COL_REFUND).value)
            balance = _num(ws.cell(r, COL_BALANCE).value)
            discount = _num(ws.cell(r, COL_DISCOUNT_AMT).value)
            location = _norm(ws.cell(r, COL_LOCATION).value) or "(unspecified)"
            mtype = _norm(ws.cell(r, COL_MEMBERSHIP_TYPE).value) or "(unspecified)"
            pmode = _norm(ws.cell(r, COL_PAYMENT_MODE).value) or "(unspecified)"
            pt = _norm(ws.cell(r, COL_PT).value).lower() in ("yes", "y", "true", "1")

            key = f"{name}||{gym_type}"
            c = clients.get(key)
            if c is None:
                c = ClientAggregate(name=name, gym_type=gym_type)
                clients[key] = c
            c.transaction_count += 1
            c.total_paid += amount
            c.cash_paid += cash
            c.online_paid += online
            c.discount_given += discount
            c.refund += refund
            c.outstanding += balance
            c.locations.add(location)
            c.membership_types.add(mtype)
            c.payment_modes.add(pmode)
            c.is_pt = c.is_pt or pt
            if c.first_seen is None or dop < c.first_seen:
                c.first_seen = dop
            if c.last_seen is None or dop > c.last_seen:
                c.last_seen = dop

            seg.by_payment_mode[gym_type][pmode] = seg.by_payment_mode[gym_type].get(pmode, 0.0) + amount
            seg.by_membership_type[gym_type][mtype] = seg.by_membership_type[gym_type].get(mtype, 0.0) + amount
            seg.by_location[gym_type][location] = seg.by_location[gym_type].get(location, 0.0) + amount

    seg.pt_clients = sorted(
        (c for c in clients.values() if c.is_pt),
        key=lambda c: c.total_paid,
        reverse=True,
    )
    return clients, seg
