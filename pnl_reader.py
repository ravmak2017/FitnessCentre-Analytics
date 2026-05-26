from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from data_sources import master_path as _active_master_path

# Backwards-compat constant — points at the *active* source's master at import time.
# Prefer calling `_active_master_path()` inside functions so Streamlit source-switching works.
MASTER_PATH = _active_master_path()
SHEET_NAME = "P&L"
FIRST_DATA_ROW = 5

COLS = {
    "month": 1,
    "ultra_rev": 2,
    "premium_rev": 3,
    "salary": 5,
    "electricity": 6,
    "repair": 7,
    "grocery": 8,
    "mandir": 9,
    "paytm": 10,
    "misc": 11,
    "diesel": 12,
}
EXPENSE_KEYS = ["salary", "electricity", "repair", "grocery", "mandir", "paytm", "misc", "diesel"]


@dataclass
class MonthlyPnL:
    month: datetime
    ultra_rev: float
    premium_rev: float
    salary: float
    electricity: float
    repair: float
    grocery: float
    mandir: float
    paytm: float
    misc: float
    diesel: float

    @property
    def total_rev(self) -> float:
        return self.ultra_rev + self.premium_rev

    @property
    def total_exp(self) -> float:
        return sum(getattr(self, k) for k in EXPENSE_KEYS)

    @property
    def profit(self) -> float:
        return self.total_rev - self.total_exp

    @property
    def gm_pct(self) -> float:
        return (self.profit / self.total_rev) if self.total_rev else 0.0

    @property
    def label(self) -> str:
        return self.month.strftime("%b %Y")


def _num(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    return 0.0


def read_pnl(path: Path | None = None) -> list[MonthlyPnL]:
    if path is None:
        path = _active_master_path()
    wb = load_workbook(path, data_only=False)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{SHEET_NAME}' not found in {path}")
    ws = wb[SHEET_NAME]

    months: list[MonthlyPnL] = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        month_val = ws.cell(r, COLS["month"]).value
        if not isinstance(month_val, datetime):
            continue
        ultra = _num(ws.cell(r, COLS["ultra_rev"]).value)
        premium = _num(ws.cell(r, COLS["premium_rev"]).value)
        if ultra == 0 and premium == 0:
            continue
        months.append(MonthlyPnL(
            month=month_val,
            ultra_rev=ultra,
            premium_rev=premium,
            salary=_num(ws.cell(r, COLS["salary"]).value),
            electricity=_num(ws.cell(r, COLS["electricity"]).value),
            repair=_num(ws.cell(r, COLS["repair"]).value),
            grocery=_num(ws.cell(r, COLS["grocery"]).value),
            mandir=_num(ws.cell(r, COLS["mandir"]).value),
            paytm=_num(ws.cell(r, COLS["paytm"]).value),
            misc=_num(ws.cell(r, COLS["misc"]).value),
            diesel=_num(ws.cell(r, COLS["diesel"]).value),
        ))
    months.sort(key=lambda m: m.month)
    return months
